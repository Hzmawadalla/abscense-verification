"""Parse the HC + Structure tabs into managers, employees, and a mapping between them (SPEC §6.1).

Pure function over a workbook path — no database. Returns structured data the seed step upserts,
and an exceptions list for anything that can't be cleanly mapped (never silently dropped).

CRM is the join key across tabs but its casing is inconsistent between HC and Structure, so all
matching is case-insensitive on a lowercased key; the canonical display casing comes from HC."""
from dataclasses import dataclass, field

from .workbook import norm_header, pick, read_dicts, to_date

PLACEHOLDERS = {"boot camp"}
_JUNK_CRM = {"n/a", "#n/a", "#ref!", "0", "na", "none"}


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _key(crm):
    """Case-insensitive join key for a cleaned CRM."""
    return crm.lower() if crm else None


def _is_junk_crm(crm) -> bool:
    """Reject malformed CRM keys (blank, N/A, numeric-only, team codes, formula errors)."""
    if crm is None:
        return True
    s = str(crm).strip().lower()
    if s in _JUNK_CRM or s == "":
        return True
    if s.replace(".", "", 1).isdigit():
        return True
    if "大组" in s or "#ref" in s or "#n/a" in s:
        return True
    return False


@dataclass
class Manager:
    crm: str
    name: str | None = None
    email: str | None = None
    active: bool = True


@dataclass
class Employee:
    crm: str
    ps_id: str | None = None
    name: str | None = None
    email: str | None = None
    department: str | None = None
    vendor: str | None = None
    employee_status: str | None = None
    join_date: object = None
    exit_date: object = None
    manager_crm: str | None = None
    sm_crm: str | None = None
    team: str | None = None


@dataclass
class RefException:
    crm: str | None
    reason: str
    detail: str | None = None


@dataclass
class ReferenceData:
    managers: list = field(default_factory=list)
    employees: list = field(default_factory=list)
    exceptions: list = field(default_factory=list)
    stats: dict = field(default_factory=dict)


def _structure_map(struct_rows):
    """employee_key -> {tl_key, tl_raw, sm_raw, team, emp_raw}."""
    mapping = {}
    for row in struct_rows:
        ecrm = _clean(pick(row, "crm"))
        if not ecrm or _is_junk_crm(ecrm):
            continue
        tl = _clean(pick(row, "tl/coach lead", "tl"))
        if tl and (tl.lower() in PLACEHOLDERS or _is_junk_crm(tl)):
            tl = None
        sm = _clean(pick(row, "sm/ltl/stl", "sm"))
        if sm and (sm.lower() in PLACEHOLDERS or _is_junk_crm(sm)):
            sm = None
        mapping[_key(ecrm)] = {
            "tl_key": _key(tl), "tl_raw": tl,
            "sm_raw": sm, "team": _clean(pick(row, "team")), "emp_raw": ecrm,
        }
    return mapping


def _find_active_sheet(wb):
    """Locate a Line-Manager export sheet and its header row: the first sheet whose header row
    (within the first few rows) carries both a 'crm account' and a 'line manager' column. Returns
    (sheet_name, header_row) or None. Covers the standalone 'All Active Employees' export (headers
    on row 2, system codes on row 1) and a combined workbook whose Line-Manager data sits on a tab
    named e.g. 'Structure' with headers on row 1."""
    for name in wb.sheetnames:
        for idx, row in enumerate(wb[name].iter_rows(min_row=1, max_row=6, values_only=True), start=1):
            hdrs = {norm_header(h) for h in row}
            if "crm account" in hdrs and ("line manager email" in hdrs or "line manager" in hdrs):
                return name, idx
    return None


def parse_active_employees(path, sheet=None, header_row=None):
    """Parse a Line-Manager export (the 'All Active Employees' style) into ReferenceData.

    Each row is an employee carrying their Line Manager. The verifier is that Line Manager,
    identified by EMAIL so a manager who is not themselves a row in the file is still a valid
    verifier; when the Line Manager *is* an in-file employee (resolvable via 'Line Manager Employee
    ID' -> that row's 'CRM account') their own CRM is used as the stable identity instead. 'Last
    Working Day' is stored as exit_date so the summary parser can skip flagged days after it.

    sheet/header_row are auto-detected when omitted (standalone export: row 2; combined workbook:
    wherever the Line-Manager headers sit)."""
    import openpyxl
    if sheet is None or header_row is None:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet, header_row = _find_active_sheet(wb) or (wb.sheetnames[0], 2)
        finally:
            wb.close()
    _, rows = read_dicts(path, sheet, header_row=header_row)

    ref = ReferenceData()
    id_to_crm = {}   # Employee ID -> CRM account, to resolve an in-file Line Manager to their CRM
    for row in rows:
        crm = _clean(pick(row, "crm account"))
        emp_id = _clean(pick(row, "employee id"))
        if crm and not _is_junk_crm(crm) and emp_id:
            id_to_crm[emp_id] = crm

    managers = {}   # identity key (lowercased crm or email) -> Manager
    seen = set()
    for row in rows:
        crm = _clean(pick(row, "crm account"))
        if not crm or _is_junk_crm(crm) or _key(crm) in seen:
            continue
        seen.add(_key(crm))
        lm_id = _clean(pick(row, "line manager employee id"))
        # Read the manager name by exact header only: pick()'s prefix fallback would match
        # 'line manager employee id' / 'line manager email' when this cell is blank, leaking the
        # id or email in as the manager's name.
        lm_name = _clean(row.get("line manager"))
        lm_email = _clean(pick(row, "line manager email"))
        emp_email = _clean(pick(row, "email", "name email"))

        # Verifier identity: an in-file CRM (resolved via the manager's employee id) first, else
        # the manager's email — so a manager who isn't a row in the file is still a valid verifier.
        in_file_crm = id_to_crm.get(lm_id) if lm_id else None
        key = _key(in_file_crm) if in_file_crm else (_key(lm_email) if lm_email else None)
        if key and key in (_key(crm), _key(emp_email)):   # reject self-management
            key = None

        if key:
            if key not in managers:
                managers[key] = Manager(crm=in_file_crm or lm_email, name=lm_name, email=lm_email)
            manager_crm = managers[key].crm   # canonical, so all reports of one manager agree
        else:
            manager_crm = None
            ref.exceptions.append(RefException(
                crm, "unmapped_employee",
                f"line manager '{lm_name or lm_email or lm_id or '?'}' could not be resolved"))
        ref.employees.append(Employee(
            crm=crm, ps_id=_clean(pick(row, "employee id")),
            name=_clean(pick(row, "name")),
            email=emp_email,
            department=_clean(pick(row, "department")),
            employee_status=_clean(pick(row, "employee status")),
            exit_date=to_date(pick(row, "last working day")),
            manager_crm=manager_crm))

    ref.managers = list(managers.values())
    mapped = sum(1 for e in ref.employees if e.manager_crm)
    ref.stats = {"managers": len(ref.managers), "employees": len(ref.employees),
                 "mapped_employees": mapped, "exceptions": len(ref.exceptions)}
    return ref


def parse_reference_any(path, aliases=None):
    """Dispatch to the right reference parser. A Line-Manager export — a sheet carrying 'CRM account'
    + 'Line Manager' columns, whether standalone or alongside an HC tab — goes to
    parse_active_employees. Otherwise the classic HC + Structure (TL/Coach Lead) workbook goes to
    parse_reference."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        found = _find_active_sheet(wb)
    finally:
        wb.close()
    if found:
        sheet, header_row = found
        return parse_active_employees(path, sheet=sheet, header_row=header_row)
    return parse_reference(path, aliases=aliases)


def parse_reference(path, hc_sheet="HC", structure_sheet="Structure", aliases=None):
    """aliases: optional {lowercased nickname/crm -> {'email':.., 'name':..}} for TLs missing from HC."""
    aliases = {k.lower(): v for k, v in (aliases or {}).items()}
    _, hc_rows = read_dicts(path, hc_sheet, 1)
    _, st_rows = read_dicts(path, structure_sheet, 1)

    ref = ReferenceData()

    # HC master keyed case-insensitively; keep HC's canonical casing + row.
    hc_by_key = {}
    for row in hc_rows:
        crm = _clean(pick(row, "crm"))
        if crm is None:
            continue
        if _is_junk_crm(crm):
            ref.exceptions.append(RefException(crm, "invalid_crm", "malformed CRM in HC"))
            continue
        hc_by_key.setdefault(_key(crm), (crm, row))

    st_map = _structure_map(st_rows)

    # Managers = distinct Team Leaders; resolve via HC, then alias, else flag.
    tl_by_key = {}
    for m in st_map.values():
        if m["tl_key"]:
            tl_by_key.setdefault(m["tl_key"], m["tl_raw"])

    managers = {}
    for tl_key, tl_raw in sorted(tl_by_key.items()):
        if tl_key in hc_by_key:
            canon, row = hc_by_key[tl_key]
            mgr = Manager(crm=canon,
                          name=_clean(pick(row, "full name", "name")),
                          email=_clean(pick(row, "work email", "email")))
            if not mgr.email and tl_key in aliases:
                mgr.email = aliases[tl_key].get("email")
                mgr.name = mgr.name or aliases[tl_key].get("name")
            if not mgr.email:
                ref.exceptions.append(RefException(canon, "manager_no_email"))
        elif tl_key in aliases:
            mgr = Manager(crm=tl_raw,
                          name=aliases[tl_key].get("name"),
                          email=aliases[tl_key].get("email"))
        else:
            mgr = Manager(crm=tl_raw)
            ref.exceptions.append(RefException(tl_raw, "manager_not_in_hc"))
        managers[tl_key] = mgr
    ref.managers = list(managers.values())

    # Employees from HC (excluding Departed), joined to their TL via Structure.
    for key, (canon, row) in hc_by_key.items():
        status = _clean(pick(row, "employee status"))
        if status and status.lower() == "departed":
            continue
        tl_crm = sm = team = None
        if key in st_map:
            m = st_map[key]
            sm, team = m["sm_raw"], m["team"]
            if m["tl_key"] is None:
                ref.exceptions.append(RefException(canon, "unmapped_employee", "no TL in Structure"))
            else:
                mgr = managers.get(m["tl_key"])
                tl_crm = mgr.crm if mgr else None
        else:
            ref.exceptions.append(RefException(canon, "unmapped_employee", "no Structure row"))
        ref.employees.append(Employee(
            crm=canon,
            ps_id=_clean(pick(row, "ps id", "ps")),
            name=_clean(pick(row, "full name", "name")),
            email=_clean(pick(row, "work email", "email")),
            department=_clean(pick(row, "department")),
            vendor=_clean(pick(row, "vendor")),
            employee_status=status,
            join_date=to_date(pick(row, "join date")),
            exit_date=to_date(pick(row, "exit date")),
            manager_crm=tl_crm, sm_crm=sm, team=team,
        ))

    # People who appear in Structure but have no HC record.
    for key, m in st_map.items():
        if key not in hc_by_key:
            ref.exceptions.append(RefException(m["emp_raw"], "employee_not_in_hc", "in Structure, not in HC"))

    ref.stats = {
        "managers": len(ref.managers),
        "employees": len(ref.employees),
        "mapped_employees": sum(1 for e in ref.employees if e.manager_crm),
        "exceptions": len(ref.exceptions),
    }
    return ref

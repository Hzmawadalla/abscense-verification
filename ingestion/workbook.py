"""Low-level workbook helpers shared by the ingestion parsers.

Reads a sheet into dicts keyed by normalized header, tolerant of the messy real-world headers
in the attendance workbook (newlines, trailing spaces, mixed case)."""
import datetime

import openpyxl

_EXCEL_EPOCH = datetime.datetime(1899, 12, 30)


def norm_header(h) -> str:
    if h is None:
        return ""
    return " ".join(str(h).replace("\n", " ").split()).strip().lower()


def resolve_sheet(wb, name):
    """Return the actual worksheet name matching `name`, tolerating surrounding whitespace and
    case (real HR exports often ship a tab like 'Structure ' with a trailing space)."""
    if name in wb.sheetnames:
        return name
    target = str(name).strip().lower()
    for s in wb.sheetnames:
        if s.strip().lower() == target:
            return s
    raise KeyError(f"Worksheet {name} does not exist.")


def read_dicts(path, sheet, header_row=1):
    """Return (headers, rows) where each row is a dict keyed by normalized header (first wins)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[resolve_sheet(wb, sheet)]
        raw = list(ws.iter_rows(min_row=1, values_only=True))
    finally:
        wb.close()
    if len(raw) < header_row:
        return [], []
    headers = [norm_header(h) for h in raw[header_row - 1]]
    rows = []
    for r in raw[header_row:]:
        d = {}
        for i, key in enumerate(headers):
            if key and key not in d:
                d[key] = r[i] if i < len(r) else None
        rows.append(d)
    return headers, rows


def pick(d, *aliases):
    """Value for the first alias present (exact match first, then header-prefix fallback)."""
    for a in aliases:
        if d.get(a) not in (None, ""):
            return d[a]
    for a in aliases:
        for k, v in d.items():
            if k.startswith(a) and v not in (None, ""):
                return v
    return None


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)):
        try:
            return (_EXCEL_EPOCH + datetime.timedelta(days=float(v))).date()
        except (ValueError, OverflowError):
            return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

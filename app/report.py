"""Build the reconciled attendance report (design 2026-07-08).

Given the original attendance workbook (the wide Summary Report matrix) and the set of closed
verification cases, produce a two-sheet .xlsx returned as bytes:

  * "<matrix sheet>" — the original matrix with each closed case's cell overwritten by its final
    verdict label, matched by CRM x date.
  * "Changes" — one row per closed case: CRM, Employee, Date, Before, After, TL, Closed by, Comment.

Matching mirrors ingestion exactly (case-insensitive CRM key, same date-header parsing), so a cell
verified during ingestion maps back to the same cell here.
"""
import io

import openpyxl

from ingestion.reference import _clean, _key
from ingestion.summary import parse_day_header
from ingestion.workbook import norm_header

CHANGES_SHEET = "Changes"
MATRIX_SHEET_HINT = "Summary Report"
CHANGES_HEADER = ["CRM", "Employee", "Date", "Before", "After", "TL", "Closed by", "Comment"]


def _locate_header(ws):
    """Return (header_row_1based, crm_col_1based, header_values) or (None, None, None).

    The header row varies by export, so it is found as the first row containing a 'CRM' cell.
    """
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        crm_col = next((j for j, h in enumerate(row) if norm_header(h) == "crm"), None)
        if crm_col is not None:
            return r_idx, crm_col + 1, row
    return None, None, None


def build_reconciled_report(matrix_path, closed_cases, labels, year,
                            sheet=MATRIX_SHEET_HINT) -> bytes:
    """closed_cases: iterable of dicts with keys employee_crm, employee_name, manager_name,
    work_date (date), source_status, final_status (verdict code), closed_by, manager_comment.
    labels: {verdict_code: human_label}. Returns the .xlsx as bytes."""
    closed_cases = list(closed_cases)
    wb = openpyxl.load_workbook(matrix_path)  # writable (not read_only) so cells can be overwritten
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    hdr_row, crm_col, header_vals = _locate_header(ws)
    if hdr_row is None:
        raise ValueError("no 'CRM' column found in the attendance matrix")

    date_cols = {}  # date -> 1-based column
    for j, h in enumerate(header_vals):
        d = parse_day_header(h, year)
        if d is not None:
            date_cols[d] = j + 1

    final_by_cell = {}  # (crm_key, date) -> final verdict code
    for cs in closed_cases:
        key = _key(_clean(cs["employee_crm"]))
        if key and cs.get("work_date") is not None:
            final_by_cell[(key, cs["work_date"])] = cs["final_status"]

    for r in range(hdr_row + 1, ws.max_row + 1):
        key = _key(_clean(ws.cell(row=r, column=crm_col).value))
        if not key:
            continue
        for d, col in date_cols.items():
            code = final_by_cell.get((key, d))
            if code is not None:
                ws.cell(row=r, column=col).value = labels.get(code, code)

    ch = wb.create_sheet(CHANGES_SHEET)
    ch.append(CHANGES_HEADER)
    for cs in closed_cases:
        wd = cs.get("work_date")
        ch.append([
            cs.get("employee_crm"),
            cs.get("employee_name"),
            wd.isoformat() if wd is not None else None,
            cs.get("source_status"),
            labels.get(cs.get("final_status"), cs.get("final_status")),
            cs.get("manager_name"),
            cs.get("closed_by"),
            cs.get("manager_comment"),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

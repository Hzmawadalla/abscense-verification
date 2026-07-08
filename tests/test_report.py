"""Behavior contract for build_reconciled_report."""
import datetime
import io

import openpyxl

from app.report import build_reconciled_report

LABELS = {"absent": "Absent", "annual_leave": "Annual Leave", "present": "Present"}


def _make_matrix(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary Report"
    ws.append(["CRM", "Normal Days", "15-Jun", "16-Jun"])
    ws.append(["51ahmed", 0, "Absent", "Normal"])
    ws.append(["51sara", 0, "Normal", "Absent"])
    p = tmp_path / "matrix.xlsx"
    wb.save(p)
    return str(p)


def _closed():
    return [{
        "employee_crm": "51AHMED",          # different casing on purpose
        "employee_name": "Ahmed Ali",
        "manager_name": "Zimmy",
        "work_date": datetime.date(2026, 6, 15),
        "source_status": "Absent",
        "final_status": "annual_leave",
        "closed_by": "hrbp",
        "manager_comment": "approved leave",
    }]


def test_overwrites_matched_cell_case_insensitive(tmp_path):
    data = build_reconciled_report(_make_matrix(tmp_path), _closed(), LABELS, year=2026)
    ws = openpyxl.load_workbook(io.BytesIO(data))["Summary Report"]
    # 51ahmed / 15-Jun (row 2, col 3) overwritten despite CRM casing mismatch
    assert ws.cell(row=2, column=3).value == "Annual Leave"


def test_leaves_untouched_cells_unchanged(tmp_path):
    data = build_reconciled_report(_make_matrix(tmp_path), _closed(), LABELS, year=2026)
    ws = openpyxl.load_workbook(io.BytesIO(data))["Summary Report"]
    assert ws.cell(row=2, column=4).value == "Normal"   # ahmed 16-Jun (not a case)
    assert ws.cell(row=3, column=3).value == "Normal"   # sara 15-Jun (different employee)


def test_changes_sheet_records_before_and_after(tmp_path):
    data = build_reconciled_report(_make_matrix(tmp_path), _closed(), LABELS, year=2026)
    rows = list(openpyxl.load_workbook(io.BytesIO(data))["Changes"].iter_rows(values_only=True))
    assert rows[0] == ("CRM", "Employee", "Date", "Before", "After", "TL", "Closed by", "Comment",
                       "In workbook?")
    assert rows[1][0] == "51AHMED"
    assert rows[1][2] == "2026-06-15"
    assert rows[1][3] == "Absent"          # before
    assert rows[1][4] == "Annual Leave"    # after
    assert rows[1][8] == "Yes"             # present in this workbook


def test_case_not_in_workbook_is_flagged_no_and_not_written(tmp_path):
    # A closed case for an employee/date absent from the uploaded file (like her 9-Jun case).
    out_of_file = [{
        "employee_crm": "EGLP-esraamahmoud",
        "employee_name": "Esraa",
        "manager_name": "Zimmy",
        "work_date": datetime.date(2026, 6, 9),   # not a column in the test matrix
        "source_status": "Bereavement Leave",
        "final_status": "present",
        "closed_by": "hrbp",
        "manager_comment": "",
    }]
    data = build_reconciled_report(_make_matrix(tmp_path), out_of_file, LABELS, year=2026)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    # nothing overwritten in the matrix (both existing rows keep their original values)
    ws = wb["Summary Report"]
    assert ws.cell(row=2, column=3).value == "Absent"
    assert ws.cell(row=3, column=3).value == "Normal"
    # she still appears in Changes, flagged "No"
    changes = list(wb["Changes"].iter_rows(values_only=True))
    assert changes[1][0] == "EGLP-esraamahmoud"
    assert changes[1][8] == "No"


def test_missing_crm_column_raises(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.append(["Employee", "15-Jun"])
    p = tmp_path / "bad.xlsx"
    wb.save(p)
    try:
        build_reconciled_report(str(p), _closed(), LABELS, year=2026)
        assert False, "expected ValueError"
    except ValueError:
        pass

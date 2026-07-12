"""Shared fixtures. Builds a synthetic workbook mirroring the real HC + Structure + Summary Report
tabs (no PII), crafted to exercise every mapping and classification edge."""
import datetime
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

HC_HEADERS = ["Type", "Vendor", "PS ID", "CRM", "Full Name", "Department",
              "Join Date\n(yyyy/mm/dd)", "Employee Status", "Exit Date\nyyyy/mm/dd",
              "Work Email address"]

HC_ROWS = [
    ["Full time", "JHR",     "1001", "TL-A", "Alice TL",  "EA", datetime.datetime(2024, 1, 1), "Active",   None,                          "alice@x.com"],
    ["Full time", "JHR",     "1002", "E-1",  "Emp One",   "EA", datetime.datetime(2024, 2, 1), "Active",   None,                          "e1@x.com"],
    ["Full time", "Migrate", "1003", "E-2",  "Emp Two",   "CC", datetime.datetime(2024, 3, 1), "Active",   None,                          "e2@x.com"],
    ["Full time", "JHR",     "1004", "E-3",  "Gone Away", "EA", datetime.datetime(2023, 1, 1), "Departed", datetime.datetime(2024, 6, 1), "e3@x.com"],
    ["Full time", "JHR",     "1005", "E-4",  "Orphan",    "EA", datetime.datetime(2024, 4, 1), "Active",   None,                          "e4@x.com"],
    ["Full time", "Migrate", "1006", "E-5",  "Fifth Emp", "CC", datetime.datetime(2024, 5, 1), "Active",   None,                          "e5@x.com"],
    ["Full time", "JHR",     "1007", "TL-C", "Carol TL",  "EA", datetime.datetime(2024, 1, 15), "Active",  None,                          "carol@x.com"],
    ["Full time", "JHR",     "1008", "E-6",  "Sixth Emp", "EA", datetime.datetime(2024, 6, 15), "Active",  None,                          "e6@x.com"],
    ["Full time", "JHR",     "1099", "N/A",  "Junk Row",  "Admin", datetime.datetime(2024, 6, 1), "Active", None,                        "junk@x.com"],
]

STRUCT_HEADERS = ["SM/LTL/STL", "TL/Coach Lead", "EA 大组Bigteam", "Team", "CRM"]
STRUCT_ROWS = [
    ["SM-A", "TL-A", "BT1", "T1", "TL-A"],   # a TL is also a member of their own team
    ["SM-A", "TL-A", "BT1", "T1", "E-1"],
    ["SM-A", "TL-A", "BT1", "T1", "E-2"],
    ["SM-A", "TL-A", "BT1", "T1", "E-3"],     # departed — should be excluded from employees
    ["boot camp", "boot camp", "boot camp", "boot camp", "E-BC"],  # placeholder + not in HC
    ["SM-B", "TL-B", "BT2", "T2", "E-5"],     # TL-B is not in HC
    ["SM-A", "tl-c", "BT1", "T3", "TL-C"],    # TL ref uses different casing than HC ('TL-C')
    ["SM-A", "tl-c", "BT1", "T3", "E-6"],     # employee under the lowercased TL ref
    # E-4 intentionally absent from Structure -> unmapped
]

# Summary Report matrix: row1 title, row2 blank, row3 headers, row4+ data.
SUMMARY_HEADERS = ["CRM", "Normal Days", "Abnormal Days", "06-May", "07-May", "08-May"]
SUMMARY_ROWS = [
    ["E-1",       20, 3, "Absent",           "Normal",                "Annual Leave"],
    ["E-2",       22, 1, "Weird Status XYZ", "Annual Leave (Failed)", "Unpaid Leave (HD) - To Be Confirmed"],
    ["E-4",       18, 5, "Absent",           "Weekend",               "Normal"],
    ["E-UNKNOWN", 10, 2, "Absent",           "Normal",                "Normal"],
]


# 'All Active Employees' single-sheet export: row 1 = system codes, row 2 = human headers,
# each row an employee carrying their own Line Manager (resolved to a CRM via Employee ID).
ACTIVE_CODES = ["sys_crm", "sys_empid", "sys_name", "sys_email", "sys_dept",
                "sys_status", "sys_lwd", "sys_lm", "sys_lmid", "sys_lmemail"]
ACTIVE_HEADERS = ["CRM account", "Employee ID", "Name", "Email", "Department",
                  "Employee Status", "Last Working Day", "Line Manager",
                  "Line Manager Employee ID", "Line Manager Email"]
ACTIVE_ROWS = [
    # A top manager: no line manager of their own -> unmapped as an employee, but verifies others.
    ["MGR-1", "9001", "Manager One", "mgr1@x.com", "EA", "Active", None, None, None, None],
    ["A-1", "1001", "Alpha", "a1@x.com", "EA", "Active", None,
     "Manager One", "9001", "mgr1@x.com"],
    ["A-2", "1002", "Beta",  "a2@x.com", "EA", "Active", datetime.datetime(2026, 5, 7),
     "Manager One", "9001", "mgr1@x.com"],
    # Line manager (id 8888) is not present in the file -> employee cannot be mapped.
    ["A-3", "1003", "Gamma", "a3@x.com", "CC", "Active", None,
     "Ghost Mgr", "8888", "ghost@x.com"],
    # Junk CRM -> skipped entirely (not an employee, not a manager).
    ["N/A", "1099", "Junk",  "j@x.com",  "Admin", "Active", None,
     "Manager One", "9001", "mgr1@x.com"],
]

# Summary matrix for the active-format fixture: row1 title, row2 blank, row3 headers, row4+ data.
ACTIVE_SUMMARY_HEADERS = ["CRM", "Normal Days", "Abnormal Days", "06-May", "07-May", "08-May"]
ACTIVE_SUMMARY_ROWS = [
    ["A-1", 20, 1, "Absent", "Normal", "Normal"],
    ["A-2", 18, 3, "Absent", "Absent", "Absent"],  # 08-May falls after A-2's last working day
]


# A manager referenced by an employee whose 'Line Manager' NAME cell is blank while the ID and
# email columns are populated — the exact shape that leaks an id/email in as the manager's name.
BLANK_LM_ROWS = [
    ["BOSS", "7000", "Big Boss", "boss@x.com", "EA", "Active", None, None, None, None],
    ["W-1", "2001", "Worker", "w1@x.com", "EA", "Active", None, "", "7000", "boss@x.com"],
]


def _build_active(wb, rows=ACTIVE_ROWS):
    ws = wb.active
    ws.title = "All Active Employees"
    ws.append(ACTIVE_CODES)
    ws.append(ACTIVE_HEADERS)
    for r in rows:
        ws.append(r)


@pytest.fixture
def active_employees_workbook(tmp_path):
    wb = openpyxl.Workbook()
    _build_active(wb)
    path = tmp_path / "active.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def blank_lm_name_workbook(tmp_path):
    wb = openpyxl.Workbook()
    _build_active(wb, rows=BLANK_LM_ROWS)
    path = tmp_path / "blank_lm.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def active_employees_with_summary_workbook(tmp_path):
    wb = openpyxl.Workbook()
    _build_active(wb)
    sm = wb.create_sheet("Summary Report")
    sm.append(["Enhanced Attendance Summary Report"])
    sm.append([])
    sm.append(ACTIVE_SUMMARY_HEADERS)
    for r in ACTIVE_SUMMARY_ROWS:
        sm.append(r)
    path = tmp_path / "active_summary.xlsx"
    wb.save(path)
    return str(path)


def _build_reference(wb):
    hc = wb.active
    hc.title = "HC"
    hc.append(HC_HEADERS)
    for r in HC_ROWS:
        hc.append(r)
    st = wb.create_sheet("Structure")
    st.append(STRUCT_HEADERS)
    for r in STRUCT_ROWS:
        st.append(r)


@pytest.fixture
def sample_workbook(tmp_path):
    wb = openpyxl.Workbook()
    _build_reference(wb)
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def sample_workbook_with_summary(tmp_path):
    wb = openpyxl.Workbook()
    _build_reference(wb)
    sm = wb.create_sheet("Summary Report")
    sm.append(["Enhanced Attendance Summary Report"])
    sm.append([])
    sm.append(SUMMARY_HEADERS)
    for r in SUMMARY_ROWS:
        sm.append(r)
    path = tmp_path / "sample_summary.xlsx"
    wb.save(path)
    return str(path)
